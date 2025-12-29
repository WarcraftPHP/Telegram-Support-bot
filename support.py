# ПРЕДУПРЕЖДЕНИЕ! ЛЮБИТЕЛЬСКИЙ КОД!
import telebot
from telebot.types import ReplyKeyboardMarkup, KeyboardButton, Message
import openpyxl 
import os
API_TOKEN = 'BOT_TOKEN' # Токен от бота
bot = telebot.TeleBot(API_TOKEN) # Бот использует константу ввиде токена от бота
ADMINS_FILE = r'path/to/excel.xlsx' # Путь к файлу Excel

def load_admins_from_xl(): # Загружает список администраторов из файла Excel
    if not os.path.exists(ADMINS_FILE):
        # Создать шаблон файла Excel если его нету
        wb = openpyxl.Workbook() 
        ws = wb.active # - 
        ws.title = 'Admins'
        ws.append(['admin_id'])
        wb.save(ADMINS_FILE) # - Сохранение файла в указанном пути файла
        print(f"Created template {ADMINS_FILE}. Please add admin IDs and restart.") # Шаблон создан, надо перезапустить программу
        return [] # Возвращает в виде списка
    
    wb = openpyxl.load_workbook(ADMINS_FILE) # Загружает Excel файл, обращаясь к пути к файлу 
    if 'Admins' not in wb.sheetnames:
        raise Exception(f"'Admins' sheet missing in {ADMINS_FILE}") # Если заголовок "Admins" отсутствует, то программа сообщает об этом
    ws = wb['Admins'] # Превращает заголовок в список
    
    admins = [] # - Создание списка администраторов
    for row in ws.iter_rows(min_row=2, values_only=True):  # Пропускает строку с заголовком, принимая только ID администраторов
        admin_id = row[0] # Учитывает ввиде списка первую строку
        if admin_id is not None:
            try:
                admins.append(int(admin_id)) # Программа обращается к списку ID администраторов
            except ValueError:
                print(f"Invalid admin_id in Excel: {admin_id}, skipping.") # Если ID администратора в списке неверный - его программа игнорирует
    return admins # возвращает переменную администраторов

ADMINS = load_admins_from_xl() # загружает список администраторов из файла
# Временный словарь, чтобы оставалась связь с админинистраторy <-> пользователем
user_to_admin = {}
admin_to_user = {}
@bot.message_handler(commands=['ID'])
def admin_id(message):
    admn_ID = message.from_user.id
    bot.send_message(message.chat.id, f"Ваш ID:{admn_ID}")
@bot.message_handler(commands=['stop'])
def handle_stop_command(message: Message):
    admin_id = message.chat.id
    if admin_id not in ADMINS:
        # Неадмины не могут использовать /stop
        bot.send_message(admin_id, "Эта команда доступна только администраторам.")
        return

    if admin_id not in admin_to_user:
        bot.send_message(admin_id, "Нет активной сессии для завершения.")
        return

    user_id = admin_to_user.pop(admin_id)
    user_to_admin.pop(user_id, None)

    # Уведомляем пользователя о завершении сессии
    bot.send_message(user_id, "Ваша сессия с оператором завершена.")
    # Уведомляем администратора
    bot.send_message(admin_id, "Вы завершили сессию с клиентом.")
def reply_keyboard():
    markup = ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    markup.add(KeyboardButton('/stop'))
    return markup
    
# Назначить пользователя к клиенту
def assign_admin(user_id):
    for admin_id in ADMINS:
        if admin_id not in admin_to_user:
            user_to_admin[user_id] = admin_id
            admin_to_user[admin_id] = user_id
            return admin_id
    return None 
# Неиспользованная часть кода !
# -----------
# Handle message from users

    # admin_id = message.chat.id
    # if admin_id in admin_to_user:
    #     user_id = admin_to_user.pop(admin_id)
    #     user_to_admin.pop(user_id, None)
    #     bot.send_message(admin_id, "Связь с клиентом закрыта")
    #     bot.send_message(user_id, "Ваша сессия с оператором была закрыта")
    # else:
    #     bot.send_message(admin_id, "Активные сессии для отмены отсуствуют")

admin_keyboard_sent = set()  # хранит ID админов, которым уже отправлен панель управления

# хендлер для пользователей
@bot.message_handler(content_types=['text', 'photo'], func=lambda message: message.chat.id not in ADMINS)
def handle_user_message(message: Message):
    user_id = message.chat.id
    username = message.from_user.username
    #  Назначаем админимтратора к пользователю
    if user_id in user_to_admin:
        admin_id = user_to_admin[user_id]
    else:
        admin_id = assign_admin(user_id)
        if not admin_id:
            bot.send_message(user_id, "Наши операторы в данный момент заняты, свяжитесь с нами позже")
            return

    # Отправляем панель администратору, если еще не отправлена
    if admin_id not in admin_keyboard_sent:
        markup = reply_keyboard()
        bot.send_message(admin_id, "Для завершения введите /stop", reply_markup=markup)
        admin_keyboard_sent.add(admin_id)

    # Если это текст
    if message.content_type == 'text':
        bot.send_message(admin_id, f"Сообщение от {username}:\n{message.text}")

    # Если это изображение
    elif message.content_type == 'photo':
        file_id = message.photo[-1].file_id  # берем оригинальное качество
        caption = f"Изображение от {username}"
        if message.caption:
            caption += f"\nПодпись: {message.caption}"
        bot.send_photo(admin_id, file_id, caption=caption)

    # Если отправил другое расширение файла
    else:
        bot.send_message(admin_id, f"Пользователь {user_id} отправил контент типа: {message.content_type}")


# Обработка сообщении от оператора
@bot.message_handler(content_types=['text', 'photo'], func=lambda message: message.chat.id in ADMINS)
def handle_admin_message(message: Message):
    admin_id = message.chat.id
    admin_username = "Оператор"
    
    if admin_id not in admin_to_user:
        bot.send_message(admin_id, "Нет доступного клиента в данный момент")
        return

    user_id = admin_to_user[admin_id]

    if message.content_type == 'text':
        # Команда /stop завершает сессию с клиентом
        if message.text == "/stop":
            return
        bot.send_message(user_id, f"{admin_username}: {message.text}")

    elif message.content_type == 'photo':
        file_id = message.photo[-1].file_id # Учитывает оригинальное качество изображения
        caption = f"{admin_username} (фото)"
        if message.caption:
            caption += f"\n{message.caption}"
        
        bot.send_photo(user_id, file_id, caption=caption)
# Носок
bot.infinity_polling()